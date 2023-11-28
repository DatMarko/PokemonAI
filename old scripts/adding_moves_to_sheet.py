
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import requests
import openpyxl
from openpyxl import load_workbook
from urllib.request import urlopen
import xlsxwriter
import xlwt
import xlrd

path1 = 'Pokemon_Dex - Copy.xlsx'
book1 = openpyxl.load_workbook(path1)

path2 = 'Pokemon_Dex.xlsx'
book2 = openpyxl.load_workbook(path2)

sheet1 = book1.active
sheet2 = book2.active

#Pokemon names from Dex - Copy
x1 = sheet1.cell(row=2, column=1)
#Pokemon names from Dex
x2 = sheet2.cell(row=3, column=2)

#counter for the for loop. starts at 2 since first two rows don't count
k = 0
#i is every pokemon name in Dex
for i in sheet2['B']:
    m = 0
    for j in sheet1['A']:
        if (i.value == j.value):
            print("Copy Pokemon Name:",j.value, "m:",m+1, "Dex Pokemon name:", i.value, "k:", k+1)
            sheet2.cell(row=k+1, column=23).value = sheet1['B'][m].value   
        m = m + 1    
    k = k + 1
book2.save(path2)