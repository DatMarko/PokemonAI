#Packages
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import requests
import openpyxl
from openpyxl import load_workbook
from urllib.request import urlopen
from geopy.geocoders import Nominatim
import xlsxwriter
import xlwt
import xlrd


#Set-up
url = "https://bulbapedia.bulbagarden.net/"
pageurl = url + "wiki/List_of_Pokémon_by_National_Pokédex_number"
response = requests.get(pageurl)
# Iterate the rows and columns


pokeList = []
poke_names = []


page = response.text
soup = BeautifulSoup(page.content, 'html.parser')

#looks at page and adds all divs that have tables and are aligned at center to a list called all_matches
all_matches = soup.find_all('table', attrs={'align':['center']})
            
#looks at every item in all_matches and            
#gets all pokemon names and adds it to poke_names list
for i in all_matches:
    list = ([a.attrs.get('href') for a in soup.select('table[align="center"] td a[title*="Pok"]')])
    #print(list[180][6:-15])
    j = 0
    for x in list:
        if str(list[j][6:-15]) not in poke_names:
            poke_names.append(str(list[j][6:-15]))
        j = j + 1

all_pokemon = []
#print(poke_names)
#getting the pokemon moves for every pokemon
k = 0

print(poke_names)
print(len(poke_names))

path = 'Pokemon_Dex - Copy.xlsx'
wb = xlrd.open_workbook(path)
book = openpyxl.Workbook()

for x in poke_names:
    new_url = url + "wiki/" + x + "_(Pokémon)"
    #print(x)
    new_response = requests.get(new_url)
    new_page = new_response.text
    poke_moves = ""
    new_soup = BeautifulSoup(new_page, 'html.parser')
    new_all_matches = new_soup.find_all('table', attrs={'roundy'})
    seen = set(poke_moves)

    #finds all moves
    for i in new_all_matches:
        list = ([a.attrs.get('href') for a in new_soup.select('table[class="roundy"] td a[title*="move"]')])
        #print(list[0])
        j = 0
        for y in list:
            if str(list[j][6:-7]) not in seen:
                seen.add(str(list[j][6:-7]))
                poke_moves = poke_moves + str(list[j][6:-7]) + ", "
            j = j + 1
    all_pokemon.append(x)
        
    #Add pokemon moves to the sheet
    row = k+2
    col = 23
    print(poke_moves)

    #inserts data of each pokemon and their moves into Dex - Copy   
    sheet = book.active
    print(row)
    sheet.cell(row=row, column=1).value = x
    sheet.cell(row=row, column=2).value = poke_moves
 
   
    #ws.write(row, col, poke_moves)
    print(x + " knows: " + str(len(poke_moves)) + " moves")
    k = k + 1

 # save the file
book.save(path)
print(all_pokemon)

'''
# i is every table of each generation
for i in all_matches:
    list = ([a.attrs.get('href') for a in soup.select('table[align="center"] td a[title*="Pok"]')])
    for x in list:
        pokeList.append(url + x)
'''
'''
#for x in pokeList:
    #print(x)

# Start of Script
no = []
name = []
generation = []
abil = []
habil = []
primary = []
secondary = []
hp = []
atk = []
defense = []
spatk = []
spdef = []
spd = []
bst = []

timer = 1

#Scraping Bulbapedia
for x in pokeList[:5]: #Pokemon total 953
    p_url = x
    response = requests.get(p_url)
    p_page = response.text
    p_soup = BeautifulSoup(p_page, 'html.parser')

    #Find pokemon dex number
    no.append(p_soup.th.big.a.span.text)

    #Find pokemon name
    pokemon_table = p_soup.find('div', class_='mw-parser-output')
    #print(pokemon_table.table.tbody.tr.text)
    #print(p_soup.th.big.a.span.text)
    #name.append(p_soup.div.p.b.text)
    #print(str(timer) + " " + (p_soup.div.p.b.text))
    timer = timer + 1

    #Find Generation
    g = p_soup.select( 'ul li span a[class*="external text"] ' )
    g = BeautifulSoup(str(g)).get_text()
    g = g[:-1][1:]
    g = g.split(",")
    generation.append(g[0])

    #Find typing
    t = p_soup.select( 'td a[href*="(type)"] span b' )
    t = BeautifulSoup(str(t)).get_text()
    t = t[:-1][1:]
    t = t.split(",")
    primary.append(t[0])
    secondary.append(t[1])

    
    Work-in-progress
    #Finding Default and Hidden Ability
    a = p_soup.select( 'td a[href*="(Ability)"] span' )
    a = BeautifulSoup(str(a)).get_text()
    a = a[:-1][1:]
    a = a.split(",")
    abil.append(a[0])
    habil.append(a[3])
    

    #Find stats
    stats = p_soup.findAll('th', attrs = {'style':['width:85px; padding-left:0.5em; padding-right:0.5em']})
    stats = ([x.text for x in stats])

    #Keep only the stats numbers and store into a list
    store = []
    for x in stats:
        store.append(re.findall(r'[0-9]?[0-9]?[0-9]', x))

    #Removing brackets and converting stats into integer
    holder = []
    for x in store:
        x = int((str(x))[:-2][2:])
        holder.append(x)

    #Store stats into appropriate list
    hp.append(holder[0])
    atk.append(holder[1])
    defense.append(holder[2])
    spatk.append(holder[3])
    spdef.append(holder[4])
    spd.append(holder[5])
    bst.append(holder[6])

pokemon = {'Dex No.': no,
           'Name': name,
           'Generation':generation,
           'Primary Type': primary,
           'Secondary Type': secondary,
           #'Ability': abil,
           #'Hidden Ability': habil,
           'Health':hp,
           'Attack':atk,
           'Defense':defense,
           'Sp. Attack':spatk,
           'Sp. Defense':spdef,
           'Speed':spd,
           'BST':bst
          }

#Create Dataframe
df=pd.DataFrame.from_dict(pokemon)

#Data Cleaning
df.loc[779:869,'Generation'] = 'Generation VII'
df.loc[870:951,'Generation'] = 'Generation VIII'
df.drop_duplicates()

#Write Csv
df.to_csv('bulbapedia_data.csv', index = None, header = True) 
'''